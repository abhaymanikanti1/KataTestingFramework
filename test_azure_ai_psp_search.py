"""
Azure AI KATA Search Test - PSP Mentor Queries
Tests PSP Search queries using Azure AI API with Bearer token authentication
Requires: az CLI for token generation
"""
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import json
import subprocess
import time

# ===== CONFIG =====
AZURE_AI_BASE_URL = "https://aif-kata2.services.ai.azure.com/api/projects/kataproject/openai"
API_VERSION = "2025-11-15-preview"
AGENT_NAME = "shikha-exp-v3"

# Benchmark file with PSP Search queries
BENCHMARK_FILE = "compare.xlsx"
PSP_SHEET_NAME = "PSP Mentor Prompts"  # Sheet with PSP queries

# Output files
OUTPUT_FILE = "azure_ai_psp_search_results.xlsx"

# Bearer token (will be fetched using az CLI)
BEARER_TOKEN = None


def get_bearer_token():
    """
    Get Azure bearer token using az CLI
    Command: az account get-access-token --resource https://ai.azure.com --query accessToken -o tsv
    """
    try:
        print("🔑 Fetching Azure bearer token...")
        result = subprocess.run(
            ['az', 'account', 'get-access-token', '--resource', 'https://ai.azure.com', '--query', 'accessToken', '-o', 'tsv'],
            capture_output=True,
            text=True,
            timeout=30
        )
        
        if result.returncode == 0:
            token = result.stdout.strip()
            if token:
                print("✅ Bearer token obtained successfully")
                return token
            else:
                print("❌ Token is empty")
                return None
        else:
            print(f"❌ Failed to get token: {result.stderr}")
            return None
    except subprocess.TimeoutExpired:
        print("❌ Timeout getting token")
        return None
    except FileNotFoundError:
        print("❌ Azure CLI not found. Install: https://docs.microsoft.com/cli/azure/install-azure-cli")
        return None
    except Exception as e:
        print(f"❌ Error getting token: {e}")
        return None


def create_conversation(bearer_token):
    """
    Create a new conversation with the Azure AI agent
    Returns: conversation_id or None
    """
    url = f"{AZURE_AI_BASE_URL}/conversations?api-version={API_VERSION}"
    
    headers = {
        'Authorization': f'Bearer {bearer_token}',
        'Content-Type': 'application/json',
        'User-Agent': 'KATA-Testing-Framework/1.0'
    }
    
    payload = {
        "conversation": "",
        "input": f"Hi {AGENT_NAME}",
        "stream": False,
        "agent": {
            "name": AGENT_NAME,
            "type": "agent_reference"
        }
    }
    
    try:
        response = requests.post(url, json=payload, headers=headers, timeout=30)
        
        if response.status_code == 200 or response.status_code == 201:
            data = response.json()
            # Extract conversation ID from response
            conversation_id = data.get('conversation', data.get('id', data.get('conversation_id')))
            return conversation_id
        else:
            print(f"  ⚠️ Failed to create conversation: {response.status_code}")
            print(f"  Response: {response.text[:200]}")
            return None
            
    except Exception as e:
        print(f"  ❌ Error creating conversation: {e}")
        return None


def query_agent(bearer_token, conversation_id, question):
    """
    Send a question to the Azure AI agent
    
    Args:
        bearer_token (str): Azure bearer token
        conversation_id (str): Conversation ID from first request
        question (str): Question to ask
        
    Returns:
        dict: Response with status, answer, sources, etc.
    """
    url = f"{AZURE_AI_BASE_URL}/responses?api-version={API_VERSION}"
    
    headers = {
        'Authorization': f'Bearer {bearer_token}',
        'Content-Type': 'application/json',
        'User-Agent': 'KATA-Testing-Framework/1.0'
    }
    
    payload = {
        "conversation": conversation_id,
        "input": question,
        "stream": False,
        "agent": {
            "name": AGENT_NAME,
            "type": "agent_reference"
        }
    }
    
    try:
        response = requests.post(url, json=payload, headers=headers, timeout=60)
        
        if response.status_code == 200:
            data = response.json()
            
            # Extract answer and sources from response
            answer = data.get('output', data.get('answer', data.get('response', str(data))))
            sources = data.get('sources', data.get('citations', []))
            
            return {
                'status': 'success',
                'status_code': 200,
                'response': answer,
                'sources': sources if isinstance(sources, list) else [],
                'metadata': data,
                'conversation_id': conversation_id
            }
        else:
            return {
                'status': 'error',
                'status_code': response.status_code,
                'response': f"HTTP {response.status_code}: {response.text[:500]}",
                'sources': [],
                'metadata': {'error': response.text},
                'conversation_id': conversation_id
            }
            
    except requests.Timeout:
        return {
            'status': 'error',
            'status_code': None,
            'response': 'Request timeout (60s)',
            'sources': [],
            'metadata': {},
            'conversation_id': conversation_id
        }
    except Exception as e:
        return {
            'status': 'error',
            'status_code': None,
            'response': f"Exception: {str(e)}",
            'sources': [],
            'metadata': {'exception': str(e)},
            'conversation_id': conversation_id
        }


def load_psp_queries_from_benchmark():
    """
    Load PSP Search queries from the benchmark Excel file
    Returns: list of questions
    """
    try:
        print(f"📂 Loading queries from {BENCHMARK_FILE}...")
        wb = openpyxl.load_workbook(BENCHMARK_FILE, data_only=True)
        
        if PSP_SHEET_NAME not in wb.sheetnames:
            print(f"❌ Sheet '{PSP_SHEET_NAME}' not found in {BENCHMARK_FILE}")
            print(f"   Available sheets: {wb.sheetnames}")
            return []
        
        ws = wb[PSP_SHEET_NAME]
        queries = []
        
        # Find the "Prompt" column (usually column B or C)
        header_row = 1
        prompt_col = None
        
        for col in range(1, 10):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value and 'prompt' in str(cell_value).lower():
                prompt_col = col
                break
        
        if not prompt_col:
            print("❌ Could not find 'Prompt' column in the sheet")
            return []
        
        print(f"✅ Found Prompt column at index {prompt_col}")
        
        # Extract questions (skip header row)
        for row in range(header_row + 1, ws.max_row + 1):
            prompt = ws.cell(row=row, column=prompt_col).value
            if prompt and str(prompt).strip():
                queries.append({
                    'row': row,
                    'prompt': str(prompt).strip()
                })
        
        print(f"✅ Loaded {len(queries)} PSP Search queries")
        return queries
        
    except FileNotFoundError:
        print(f"❌ Benchmark file not found: {BENCHMARK_FILE}")
        return []
    except Exception as e:
        print(f"❌ Error loading queries: {e}")
        return []


def create_excel_report(results, bearer_token_valid):
    """
    Create an Excel file with test results
    
    Args:
        results (list): List of test result dictionaries
        bearer_token_valid (bool): Whether bearer token was valid
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PSP Search Test Results"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    success_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    error_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["#", "Benchmark Row", "Question", "Status", "Status Code", "Response", "Sources", "Conversation ID", "Timestamp"]
    ws.append(headers)
    
    # Style headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Add data
    for idx, result in enumerate(results, 1):
        sources_text = "\n".join(result.get('sources', [])) if result.get('sources') else "No sources"
        
        row_data = [
            idx,
            result.get('benchmark_row', 'N/A'),
            result['query'],
            result['status'],
            result.get('status_code', 'N/A'),
            result['response'][:5000] if result['response'] else '',  # Limit response length
            sources_text,
            result.get('conversation_id', 'N/A'),
            result['timestamp']
        ]
        
        ws.append(row_data)
        
        # Style row based on status
        row_num = idx + 1
        fill = success_fill if result['status'] == 'success' else error_fill
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = fill
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = border
    
    # Adjust column widths
    column_widths = [5, 12, 50, 10, 12, 100, 60, 40, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add summary sheet
    summary_ws = wb.create_sheet("Summary", 0)
    summary_ws["A1"] = "Azure AI PSP Search Test Summary"
    summary_ws["A1"].font = Font(bold=True, size=14)
    
    summary_ws["A3"] = "Total Tests:"
    summary_ws["B3"] = len(results)
    
    successful = sum(1 for r in results if r['status'] == 'success')
    failed = len(results) - successful
    
    summary_ws["A4"] = "Successful:"
    summary_ws["B4"] = successful
    summary_ws["B4"].fill = success_fill
    
    summary_ws["A5"] = "Failed:"
    summary_ws["B5"] = failed
    summary_ws["B5"].fill = error_fill
    
    summary_ws["A6"] = "Success Rate:"
    summary_ws["B6"] = f"{(successful/len(results)*100):.1f}%" if results else "N/A"
    
    summary_ws["A8"] = "Test Date:"
    summary_ws["B8"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    summary_ws["A9"] = "API Endpoint:"
    summary_ws["B9"] = AZURE_AI_BASE_URL
    
    summary_ws["A10"] = "Agent:"
    summary_ws["B10"] = AGENT_NAME
    
    summary_ws["A11"] = "Bearer Token Valid:"
    summary_ws["B11"] = "Yes" if bearer_token_valid else "No"
    summary_ws["B11"].fill = success_fill if bearer_token_valid else error_fill
    
    summary_ws.column_dimensions["B"].width = 80
    
    # Save workbook
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Results saved to: {OUTPUT_FILE}")


def main():
    """Main test execution function"""
    print("=" * 80)
    print("Azure AI KATA Search Test - PSP Mentor Queries")
    print("=" * 80)
    print(f"API Endpoint: {AZURE_AI_BASE_URL}")
    print(f"Agent: {AGENT_NAME}")
    print(f"Benchmark File: {BENCHMARK_FILE}")
    print(f"Output File: {OUTPUT_FILE}")
    print("=" * 80)
    print()
    
    # Get bearer token
    bearer_token = get_bearer_token()
    if not bearer_token:
        print("\n❌ Failed to obtain bearer token. Exiting.")
        print("\n💡 Make sure you're logged in to Azure CLI:")
        print("   az login")
        return
    
    print(f"Token preview: {bearer_token[:50]}...\n")
    
    # Load queries from benchmark file
    queries = load_psp_queries_from_benchmark()
    if not queries:
        print("\n❌ No queries loaded. Exiting.")
        return
    
    print(f"\n📊 Testing {len(queries)} PSP Search queries...\n")
    
    results = []
    
    for idx, query_data in enumerate(queries, 1):
        query = query_data['prompt']
        benchmark_row = query_data['row']
        
        print(f"[{idx}/{len(queries)}] Row {benchmark_row}: {query[:60]}...")
        
        # Step 1: Create conversation
        print("  → Creating conversation...")
        conversation_id = create_conversation(bearer_token)
        
        if not conversation_id:
            print("  ❌ Failed to create conversation")
            result = {
                'status': 'error',
                'status_code': None,
                'response': 'Failed to create conversation',
                'sources': [],
                'conversation_id': None,
                'query': query,
                'benchmark_row': benchmark_row,
                'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
        else:
            print(f"  ✅ Conversation created: {conversation_id}")
            
            # Small delay between requests
            time.sleep(0.5)
            
            # Step 2: Query the agent
            print("  → Querying agent...")
            result = query_agent(bearer_token, conversation_id, query)
            result['query'] = query
            result['benchmark_row'] = benchmark_row
            result['timestamp'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Print result summary
            if result['status'] == 'success':
                print(f"  ✅ Success! Response: {result['response'][:100]}...")
            else:
                print(f"  ❌ Failed: {result['response'][:100]}")
        
        results.append(result)
        print()
        
        # Small delay between tests
        time.sleep(1)
    
    # Create Excel report
    print("\nGenerating Excel report...")
    create_excel_report(results, True)
    
    # Print summary
    successful = sum(1 for r in results if r['status'] == 'success')
    failed = len(results) - successful
    
    print("\n" + "=" * 80)
    print("TEST SUMMARY")
    print("=" * 80)
    print(f"Total Tests: {len(results)}")
    print(f"✅ Successful: {successful}")
    print(f"❌ Failed: {failed}")
    print(f"Success Rate: {(successful/len(results)*100):.1f}%")
    print("=" * 80)
    print(f"\n📁 Results: {OUTPUT_FILE}")
    print("\n💡 Note: Bearer token expires after 1 hour.")
    print("   Regenerate with: az account get-access-token --resource https://ai.azure.com --query accessToken -o tsv")


if __name__ == "__main__":
    main()
