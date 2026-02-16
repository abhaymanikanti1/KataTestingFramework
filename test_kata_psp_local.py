"""
Local test for KATA PSP Search functionality using Azure AI Agent
Tests only PSP Mentor Prompts from compare.xlsx
"""
import openpyxl
import requests
import subprocess
from datetime import datetime

# Azure AI Configuration
AZURE_AI_BASE_URL = "https://aif-kata2.services.ai.azure.com/api/projects/kataproject/openai"
API_VERSION = "2025-11-15-preview"
AGENT_NAME = "shikha-exp-v3"

# Files
BENCHMARK_FILE = "compare.xlsx"
OUTPUT_FILE = f"PSP_Test_Results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

def get_bearer_token():
    """Get fresh Bearer token from Azure CLI"""
    print("🔑 Getting Bearer token from Azure CLI...")
    try:
        result = subprocess.run(
            ['az', 'account', 'get-access-token', '--resource', 'https://ai.azure.com', '--query', 'accessToken', '-o', 'tsv'],
            capture_output=True,
            text=True,
            check=True
        )
        token = result.stdout.strip()
        print(f"✅ Token obtained (expires in 1 hour)")
        return token
    except subprocess.CalledProcessError as e:
        print(f"❌ Failed to get token: {e}")
        print(f"stderr: {e.stderr}")
        raise

def create_conversation(token):
    """Step 1: Create a conversation"""
    url = f"{AZURE_AI_BASE_URL}/conversations?api-version={API_VERSION}"
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json',
        'User-Agent': 'Python-Test-Script'
    }
    payload = {
        "conversation": "",
        "input": f"Hi {AGENT_NAME}",
        "stream": False,
        "agent": {
            "name": AGENT_NAME,
            "type": "agent_reference"
        }
    }
    
    response = requests.post(url, json=payload, headers=headers, timeout=30)
    
    # Debug: print error details if request fails
    if not response.ok:
        print(f"  ⚠️  API Error {response.status_code}: {response.text[:200]}")
    
    response.raise_for_status()
    data = response.json()
    
    # Azure AI API returns the conversation ID in the 'id' field
    conversation_id = data.get('id', '')
    print(f"  ✅ Conversation ID: {conversation_id}")
    return conversation_id

def query_agent(token, conversation_id, question):
    """Step 2: Query the agent with the question"""
    url = f"{AZURE_AI_BASE_URL}/responses?api-version={API_VERSION}"
    headers = {
        'Authorization': f'Bearer {token}',
        'Content-Type': 'application/json',
        'User-Agent': 'Python-Test-Script'
    }
    payload = {
        "conversation": conversation_id,
        "input": question,
        "stream": False,
        "agent": {
            "name": AGENT_NAME,
            "type": "agent_reference"
        }
    }
    
    response = requests.post(url, json=payload, headers=headers, timeout=60)
    
    # Debug: print error details if request fails
    if not response.ok:
        print(f"  ⚠️  API Error {response.status_code}: {response.text[:200]}")
    
    response.raise_for_status()
    data = response.json()
    
    # Azure AI Agent response: output is an array of workflow actions and messages
    # We need to find the message with type='message' and role='assistant'
    output = data.get('output', [])
    answer = 'No response found'
    
    if isinstance(output, list):
        # Search for assistant messages in the output array
        for item in output:
            if isinstance(item, dict) and item.get('type') == 'message' and item.get('role') == 'assistant':
                # Found a message - extract the text from content
                content = item.get('content', [])
                if isinstance(content, list) and len(content) > 0:
                    text = content[0].get('text', '') if isinstance(content[0], dict) else str(content[0])
                    if text and text.strip():
                        answer = text.strip()
                        break
    
    print(f"  ✅ Response: {answer[:150]}...")
    return answer

def load_psp_questions():
    """Load PSP Mentor Prompts from compare.xlsx"""
    print(f"\n📂 Loading PSP questions from {BENCHMARK_FILE}...")
    wb = openpyxl.load_workbook(BENCHMARK_FILE)
    
    # PSP Mentor Prompts sheet
    sheet_name = "PSP Mentor Prompts"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {BENCHMARK_FILE}")
    
    ws = wb[sheet_name]
    questions = []
    
    # Read questions from column A (starting row 2)
    # To limit for testing, change to: max_rows = min(5, ws.max_row + 1)
    max_rows = ws.max_row + 1  # Process ALL questions
    for row in range(2, max_rows):
        question = ws.cell(row, 1).value  # Column A
        if question and str(question).strip():
            questions.append({
                'row': row,
                'question': str(question).strip()
            })
    
    print(f"✅ Loaded {len(questions)} PSP questions")
    return questions

def save_results(questions, results):
    """Save test results to Excel"""
    print(f"\n💾 Saving results to {OUTPUT_FILE}...")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PSP Test Results"
    
    # Headers
    ws['A1'] = 'Row'
    ws['B1'] = 'Question'
    ws['C1'] = 'Response'
    ws['D1'] = 'Status'
    ws['E1'] = 'Conversation ID'
    
    # Make headers bold
    from openpyxl.styles import Font
    for cell in ['A1', 'B1', 'C1', 'D1', 'E1']:
        ws[cell].font = Font(bold=True)
    
    # Data rows
    for i, (q, r) in enumerate(zip(questions, results), start=2):
        ws[f'A{i}'] = q['row']
        ws[f'B{i}'] = q['question']
        ws[f'C{i}'] = r['response']
        ws[f'D{i}'] = r['status']
        ws[f'E{i}'] = r['conversation_id'][:50] if r['conversation_id'] else ''
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 80
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 30
    
    wb.save(OUTPUT_FILE)
    print(f"✅ Results saved to {OUTPUT_FILE}")

def main():
    print("=" * 80)
    print("🧪 KATA PSP Search Agent - Local Test")
    print("=" * 80)
    
    # Get Bearer token
    token = get_bearer_token()
    
    # Load questions
    questions = load_psp_questions()
    
    print(f"\n🚀 Testing {len(questions)} PSP questions...")
    print("=" * 80)
    
    results = []
    
    for i, q in enumerate(questions, start=1):
        print(f"\n[{i}/{len(questions)}] Row {q['row']}: {q['question'][:60]}...")
        
        try:
            # Step 1: Create conversation
            conversation_id = create_conversation(token)
            
            # Step 2: Query agent
            response = query_agent(token, conversation_id, q['question'])
            
            results.append({
                'response': response,
                'status': 'SUCCESS',
                'conversation_id': conversation_id
            })
            
            print(f"  ✅ Response: {response[:100]}...")
            
        except Exception as e:
            print(f"  ❌ Error: {e}")
            results.append({
                'response': f"ERROR: {str(e)}",
                'status': 'FAILED',
                'conversation_id': ''
            })
    
    # Save results
    save_results(questions, results)
    
    # Summary
    print("\n" + "=" * 80)
    print("📊 Test Summary")
    print("=" * 80)
    successful = sum(1 for r in results if r['status'] == 'SUCCESS')
    failed = sum(1 for r in results if r['status'] == 'FAILED')
    print(f"✅ Successful: {successful}/{len(results)}")
    print(f"❌ Failed: {failed}/{len(results)}")
    print(f"📄 Results saved to: {OUTPUT_FILE}")
    print("=" * 80)

if __name__ == "__main__":
    main()
