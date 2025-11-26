"""
Integrated Testing & Comparison System
- Tests first 10 rows from each sheet (PSP, VSM, TPI, Search)
- Compares with benchmark (old) responses
- Identifies objectively worse responses
- Saves degraded responses to separate Excel
- Sends Microsoft Teams alerts for issues
"""
import openpyxl
import requests
import uuid
import urllib.parse
import os
import re
import json
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill

# ===== CONFIG =====
BENCHMARK_FILE = "compare.xlsx"  # Benchmark file with good/neutral/bad markings
NEW_OUTPUT_FILE = "Direct Query Master List V_Test.xlsx"  # New test results
DEGRADED_OUTPUT_FILE = "Degraded_Responses_Report.xlsx"  # Bad responses

# Test mode: process only first 10 rows
TEST_LIMIT = None  # Set to None for full processing

# --- Base API Configuration ---
API_BASE_URL = "https://container-app-ui-stage.purplesky-e0183d2f.eastus.azurecontainerapps.io/"
EMAIL_ID = "abhay.manikanti@fortive.com"
API_KEY = "d7e8f9b6-92a4-48e2-a0cd-f81c993f29c1"

HEADERS = {
    'Content-Type': "application/x-www-form-urlencoded",
    'User-Agent': "insomnia/11.4.0"
}

SESSION_ID = "43908e3d-7fee-4688-a6c5-f3bd32a94ffd"

# Sheet configurations
SHEET_CONFIGS = {
    3: {'name': 'PSP Mentor', 'api_path': '/api/pspmentor', 'agent_id': 'psp'},
    4: {'name': 'VSM Mentor', 'api_path': '/api/vsmmentor', 'agent_id': 'vsm'},
    5: {'name': 'TPI Mentor', 'api_path': '/api/tpimentor', 'agent_id': 'tpi'},
    6: {'name': 'Search/Chat', 'api_path': '/api/chat', 'agent_id': 'search'}
}

# Microsoft Teams Webhook URL - "Kata Bugs Webhook Channel"
TEAMS_WEBHOOK_URL = os.environ.get('TEAMS_WEBHOOK_URL', 'https://fortive.webhook.office.com/webhookb2/9021b415-1b53-4310-826c-46b64d55be8c@0f634ac3-b39f-41a6-83ba-8f107876c692/IncomingWebhook/76c795c5d84c47daa0137f5f8b4ac3f3/91b293f9-febd-40ad-ba66-36b2f7f5c200/V2pmV-S1_3evMkEnKbsv4muMg8dqENYzOyZk0R-52bibA1')

# Power Automate Workflow URL for SharePoint Upload
SHAREPOINT_UPLOAD_URL = os.environ.get('SHAREPOINT_UPLOAD_URL', 'https://default0f634ac3b39f41a683ba8f107876c6.92.environment.api.powerplatform.com:443/powerautomate/automations/direct/workflows/ba32c921bc9645aa889b6e147082b435/triggers/manual/paths/invoke?api-version=1&sp=%2Ftriggers%2Fmanual%2Frun&sv=1.0&sig=EazBMjq9CYJiZC_CU2s0n1WKCaqpZ7goHUN3LM7GzXk')

# Enable/disable Teams alerts and SharePoint upload
ENABLE_TEAMS_ALERTS = True  # Now enabled with working Office 365 Incoming Webhook!
ENABLE_SHAREPOINT_UPLOAD = True  # Upload degraded responses report to SharePoint

# ===== API FUNCTIONS =====
def send_question_to_api(prompt, api_url, agent_id):
    """Send a question to the API and return the response"""
    conversation_id = str(uuid.uuid4())
    
    payload_dict = {
        "email_id": EMAIL_ID,
        "question": prompt,
        "session_id": SESSION_ID,
        "conversation_id": conversation_id,
        "agent_id": agent_id,
        "thread_id": "",  # Empty string as per production requirements
        "selected_column": "",  # Required parameter
        "container": "useruploaded"
    }
    
    payload_encoded = urllib.parse.urlencode(payload_dict)
    
    try:
        response = requests.post(api_url, data=payload_encoded, headers=HEADERS, verify=False, timeout=60)
        
        if response.status_code == 200:
            if not response.text.strip():
                return {'status': 'success', 'response': 'Empty response from API', 'sources': []}
            
            response_text = response.text.strip()
            
            if response_text.startswith('data:') or '\ndata:' in response_text:
                return parse_sse_response(response_text)
            else:
                return parse_json_response(response_text)
        else:
            return {'status': 'error', 'response': f"HTTP {response.status_code}", 'sources': []}
            
    except requests.Timeout:
        return {'status': 'error', 'response': 'Request timeout (60s)', 'sources': []}
    except Exception as e:
        return {'status': 'error', 'response': f"Error: {str(e)[:100]}", 'sources': []}

def parse_sse_response(response_text):
    """Parse SSE format response"""
    content_parts = []
    extracted_urls = []
    
    for line in response_text.split('\n'):
        line = line.strip()
        if line.startswith('data: '):
            try:
                json_part = line[6:].strip()
                if json_part and json_part != '[DONE]':
                    data_obj = json.loads(json_part)
                    
                    if isinstance(data_obj, dict):
                        # NEW FORMAT: assistant_output directly in the object
                        if 'assistant_output' in data_obj and data_obj['assistant_output']:
                            output = str(data_obj['assistant_output'])
                            # Filter out initialization messages
                            if not any(emoji in output for emoji in ['🔄', '🔧', '📋', '🔍', '📂', '🤔', '📚', '✨']):
                                content_parts.append(output)
                        
                        # OLD FORMAT: nested under 'data' key
                        elif 'data' in data_obj and isinstance(data_obj['data'], dict):
                            chunk_data = data_obj['data']
                            
                            if 'assistant_output' in chunk_data and chunk_data['assistant_output']:
                                content_parts.append(str(chunk_data['assistant_output']))
                            elif 'content' in chunk_data and chunk_data['content']:
                                content_parts.append(str(chunk_data['content']))
                            
                            if 'sources' in chunk_data and isinstance(chunk_data['sources'], list):
                                for source in chunk_data['sources']:
                                    if isinstance(source, dict):
                                        for url_field in ['page_url', 'url', 'link', 'source_url']:
                                            if url_field in source and source[url_field]:
                                                extracted_urls.append(source[url_field])
                                                break
                        
                        # Fallback: direct content field
                        elif 'content' in data_obj and data_obj['content']:
                            content_parts.append(str(data_obj['content']))
                        
                        # Extract sources from top-level sources field
                        if 'sources' in data_obj and isinstance(data_obj['sources'], list):
                            for source in data_obj['sources']:
                                if isinstance(source, dict):
                                    for url_field in ['page_url', 'url', 'link', 'source_url']:
                                        if url_field in source and source[url_field]:
                                            extracted_urls.append(source[url_field])
                                            break
                            
            except json.JSONDecodeError:
                continue
                            
            except json.JSONDecodeError:
                continue
    
    text_content = ''.join(content_parts)
    
    if not text_content:
        text_content = "SSE response: No content extracted"
    
    if not extracted_urls and text_content:
        url_pattern = r'https?://[^\s<>"{}|\\^`\[\]]+'
        content_urls = re.findall(url_pattern, text_content + response_text)
        extracted_urls.extend(content_urls[:5])
    
    unique_urls = list(dict.fromkeys(extracted_urls))
    
    return {'status': 'success', 'response': text_content, 'sources': unique_urls[:10]}

def parse_json_response(response_text):
    """Parse JSON format response"""
    try:
        result = json.loads(response_text)
        
        text_content = ""
        if isinstance(result, dict) and "content" in result:
            text_content = str(result["content"])
        else:
            text_content = "Unexpected JSON format: " + str(result)[:200]
        
        return {'status': 'success', 'response': text_content, 'sources': []}
    except json.JSONDecodeError as e:
        return {'status': 'error', 'response': f"JSON parse error: {str(e)}", 'sources': []}

# ===== COMPARISON FUNCTIONS =====
def is_response_degraded(old_response, new_response, prompt, old_quality="unknown"):
    """
    Determine if new response is objectively worse than old response.
    Takes into account the quality marking from benchmark (good/neutral/bad).
    Returns (is_degraded: bool, reason: str, severity: str)
    """
    
    if not old_response or not new_response:
        return False, "", ""
    
    old_response = str(old_response).strip()
    new_response = str(new_response).strip()
    
    # SPECIAL CASE: If old response was marked as "bad", new response is improvement
    if old_quality == "bad":
        print(f"    ℹ️  Old response was marked BAD - any new response is potential improvement")
        return False, "", ""
    
    # Check 1: Error responses
    error_indicators = ['error', 'timeout', 'http ', 'failed', 'exception', 'empty response']
    new_has_error = any(indicator in new_response.lower() for indicator in error_indicators)
    old_has_error = any(indicator in old_response.lower() for indicator in error_indicators)
    
    if new_has_error and not old_has_error:
        severity = "HIGH" if old_quality == "good" else "MEDIUM"
        return True, "New response contains error, old response was valid", severity
    
    # Check 2: Significantly shorter response (>50% shorter)
    if len(new_response) < len(old_response) * 0.5 and len(old_response) > 100:
        # More severe if old was marked as good
        severity = "HIGH" if old_quality == "good" else "MEDIUM"
        return True, f"Response significantly shorter (Old: {len(old_response)} chars, New: {len(new_response)} chars)", severity
    
    # Check 3: Generic/unhelpful responses
    generic_responses = [
        "i don't have information",
        "i cannot help",
        "i don't know",
        "no information available",
        "unable to provide",
        "sorry, i can't",
        "i'm not sure"
    ]
    new_is_generic = any(generic in new_response.lower() for generic in generic_responses)
    old_is_generic = any(generic in old_response.lower() for generic in generic_responses)
    
    if new_is_generic and not old_is_generic and len(old_response) > 50:
        severity = "HIGH" if old_quality == "good" else "MEDIUM"
        return True, "New response is generic/unhelpful, old response was specific", severity
    
    # Check 4: Missing key terms from prompt
    prompt_lower = prompt.lower()
    prompt_keywords = [word for word in prompt_lower.split() if len(word) > 4][:5]  # First 5 significant words
    
    old_keyword_count = sum(1 for kw in prompt_keywords if kw in old_response.lower())
    new_keyword_count = sum(1 for kw in prompt_keywords if kw in new_response.lower())
    
    if new_keyword_count < old_keyword_count * 0.5 and old_keyword_count >= 2:
        severity = "MEDIUM"
        return True, f"New response less relevant (Old: {old_keyword_count} keywords, New: {new_keyword_count} keywords)", severity
    
    # Check 5: Repetitive content
    if len(new_response) > 100:
        words = new_response.split()
        unique_words = len(set(words))
        if len(words) > 0 and unique_words / len(words) < 0.3:  # Less than 30% unique words
            return True, "New response is highly repetitive", "MEDIUM"
    
    # Check 6: If old was marked as "good" and new differs significantly
    if old_quality == "good":
        # Calculate similarity (simple word overlap)
        old_words = set(old_response.lower().split())
        new_words = set(new_response.lower().split())
        
        if len(old_words) > 0:
            overlap = len(old_words & new_words) / len(old_words)
            if overlap < 0.3 and len(old_response) > 100:  # Less than 30% word overlap
                return True, "New response differs significantly from GOOD benchmark (low word overlap)", "MEDIUM"
    
    # No degradation detected
    return False, "", ""

# ===== EXCEL FUNCTIONS =====
def load_benchmark_data(sheet_number):
    """
    Load benchmark responses from compare.xlsx with quality markings.
    Scans all columns to understand structure including good/neutral/bad markings.
    """
    try:
        wb = openpyxl.load_workbook(BENCHMARK_FILE)
        sheetnames = wb.sheetnames
        
        if sheet_number < 1 or sheet_number > len(sheetnames):
            return {}
        
        ws = wb[sheetnames[sheet_number - 1]]
        
        print(f"  📊 Analyzing benchmark sheet structure...")
        
        # First, scan the header row to understand column structure
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        print(f"  📋 Columns found: {len(header_row)}")
        
        # Map column names to indices
        col_mapping = {}
        for idx, header in enumerate(header_row, 1):
            if header:
                header_lower = str(header).lower().strip()
                col_mapping[header_lower] = idx
                print(f"     Col {idx}: {header}")
        
        # Identify key columns
        prompt_col = None
        response_col = None
        sources_col = None
        quality_col = None
        rating_col = None
        
        # Look for prompt column
        for key in ['prompt', 'question', 'query']:
            if key in col_mapping:
                prompt_col = col_mapping[key]
                break
        
        # Look for response column
        for key in ['output', 'response', 'answer']:
            if key in col_mapping:
                response_col = col_mapping[key]
                break
        
        # Look for sources column
        for key in ['sources', 'source', 'urls', 'references']:
            if key in col_mapping:
                sources_col = col_mapping[key]
                break
        
        # Look for quality markings
        for key in ['quality', 'rating', 'mark', 'status', 'evaluation', 'grade']:
            if key in col_mapping:
                quality_col = col_mapping[key]
                print(f"  ⭐ Quality marking column found: Col {quality_col}")
                break
        
        # Look for additional rating column
        for key in ['good', 'bad', 'neutral', 'score']:
            if key in col_mapping:
                rating_col = col_mapping[key]
                print(f"  ⭐ Rating column found: Col {rating_col}")
                break
        
        # If no explicit column names, use positional defaults
        if not prompt_col:
            prompt_col = 2  # Column B
            print(f"  ℹ️  Using default prompt column: B (2)")
        if not response_col:
            response_col = 3  # Column C
            print(f"  ℹ️  Using default response column: C (3)")
        if not sources_col:
            sources_col = 4  # Column D
            print(f"  ℹ️  Using default sources column: D (4)")
        
        benchmark_data = {}
        
        # Load data rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 1):
            prompt = row[prompt_col - 1].value if prompt_col <= len(row) else None
            response = row[response_col - 1].value if response_col <= len(row) else None
            sources = row[sources_col - 1].value if sources_col <= len(row) else None
            
            # Get quality marking if available
            quality_mark = None
            if quality_col and quality_col <= len(row):
                quality_mark = str(row[quality_col - 1].value).lower() if row[quality_col - 1].value else None
            
            # Get rating if available
            rating = None
            if rating_col and rating_col <= len(row):
                rating = str(row[rating_col - 1].value).lower() if row[rating_col - 1].value else None
            
            # Determine overall quality
            quality_status = "unknown"
            if quality_mark:
                if any(word in quality_mark for word in ['good', 'excellent', 'pass', 'acceptable', '✓', '✔']):
                    quality_status = "good"
                elif any(word in quality_mark for word in ['bad', 'poor', 'fail', 'unacceptable', '✗', '✘']):
                    quality_status = "bad"
                elif any(word in quality_mark for word in ['neutral', 'average', 'ok', 'acceptable', '~']):
                    quality_status = "neutral"
            
            if rating:
                if any(word in rating for word in ['good', 'excellent', 'pass', 'acceptable', '✓', '✔']):
                    quality_status = "good"
                elif any(word in rating for word in ['bad', 'poor', 'fail', 'unacceptable', '✗', '✘']):
                    quality_status = "bad"
                elif any(word in rating for word in ['neutral', 'average', 'ok', 'acceptable', '~']):
                    quality_status = "neutral"
            
            if prompt:
                benchmark_data[row_idx] = {
                    'prompt': str(prompt),
                    'response': str(response) if response else "",
                    'sources': str(sources) if sources else "",
                    'quality': quality_status,
                    'quality_mark': quality_mark,
                    'rating': rating
                }
        
        # Count by quality
        quality_counts = {'good': 0, 'bad': 0, 'neutral': 0, 'unknown': 0}
        for data in benchmark_data.values():
            quality_counts[data['quality']] += 1
        
        print(f"  📚 Loaded {len(benchmark_data)} benchmark responses")
        print(f"     ✅ Good: {quality_counts['good']}")
        print(f"     ❌ Bad: {quality_counts['bad']}")
        print(f"     ⚪ Neutral: {quality_counts['neutral']}")
        print(f"     ❓ Unknown: {quality_counts['unknown']}")
        
        return benchmark_data
        
    except FileNotFoundError:
        print(f"  ❌ Benchmark file not found: {BENCHMARK_FILE}")
        return {}
    except Exception as e:
        print(f"  ❌ Error loading benchmark: {e}")
        import traceback
        traceback.print_exc()
        return {}

def create_degraded_responses_sheet(wb, sheet_name):
    """Create or get sheet for degraded responses"""
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)
    
    # Style definitions
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")  # Red
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Create headers
    headers = ["Serial Number", "Prompt", "Old Response (Benchmark)", "New Response", 
               "Old Sources", "New Sources", "Benchmark Quality", "Degradation Reason", "Severity"]
    ws.append(headers)
    
    # Style header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Set column widths with autofit
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 50
    ws.column_dimensions['I'].width = 15
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    return ws

def process_sheet_with_comparison(sheet_number, config):
    """Process sheet, compare with benchmark, identify degraded responses"""
    
    print(f"\n{'='*70}")
    print(f"📊 Processing Sheet {sheet_number}: {config['name']}")
    print(f"{'='*70}")
    
    # Load benchmark data
    benchmark_data = load_benchmark_data(sheet_number)
    
    # Load or create new workbook
    if os.path.exists(NEW_OUTPUT_FILE):
        wb_new = openpyxl.load_workbook(NEW_OUTPUT_FILE)
    else:
        wb_new = openpyxl.load_workbook(BENCHMARK_FILE)
    
    sheetnames = wb_new.sheetnames
    if sheet_number < 1 or sheet_number > len(sheetnames):
        print(f"  ❌ Invalid sheet number")
        return [], 0, 0
    
    ws_new = wb_new[sheetnames[sheet_number - 1]]
    
    # API configuration
    api_url = f"{API_BASE_URL}{config['api_path']}"
    agent_id = config['agent_id']
    
    print(f"🌐 API: {api_url}")
    print(f"🤖 Agent: {agent_id}")
    if TEST_LIMIT is None:
        print(f"⚠️  TEST MODE: Full run (no row limit)")
    else:
        print(f"⚠️  TEST MODE: First {TEST_LIMIT} rows only")
    
    processed = 0
    successful = 0
    degraded_responses = []
    
    for row_idx, row in enumerate(ws_new.iter_rows(min_row=2), 1):
        if TEST_LIMIT is not None and processed >= TEST_LIMIT:
            break
        
        prompt_cell = row[1]
        output_cell = row[2]
        sources_cell = row[3]
        
        prompt = prompt_cell.value
        if not prompt or not str(prompt).strip():
            continue
        
        processed += 1
        if TEST_LIMIT is None:
            print(f"\n[{processed}] Row {row_idx}:")
        else:
            print(f"\n[{processed}/{TEST_LIMIT}] Row {row_idx}:")
        print(f"  📝 Prompt: {str(prompt)[:60]}...")
        
        # Get benchmark response
        old_response = benchmark_data.get(row_idx, {}).get('response', '')
        old_sources = benchmark_data.get(row_idx, {}).get('sources', '')
        old_quality = benchmark_data.get(row_idx, {}).get('quality', 'unknown')
        old_quality_mark = benchmark_data.get(row_idx, {}).get('quality_mark', '')
        
        if old_quality_mark:
            print(f"     📊 Benchmark quality: {old_quality.upper()} ({old_quality_mark})")
        
        # Send to API
        result = send_question_to_api(str(prompt), api_url, agent_id)
        
        if result['status'] == 'success':
            new_response = result['response']
            new_sources = "\n".join(result['sources']) if result['sources'] else ""
            
            output_cell.value = new_response
            sources_cell.value = new_sources
            successful += 1
            
            print(f"  ✅ Success: {len(new_response)} chars")
            
            # Compare with benchmark
            is_degraded, reason, severity = is_response_degraded(old_response, new_response, str(prompt), old_quality)
            
            if is_degraded:
                print(f"  ⚠️  DEGRADATION DETECTED: {reason} (Severity: {severity})")
                
                degraded_responses.append({
                    'serial': row_idx,
                    'prompt': str(prompt),
                    'old_response': old_response,
                    'new_response': new_response,
                    'old_sources': old_sources,
                    'new_sources': new_sources,
                    'reason': reason,
                    'severity': severity,
                    'sheet_name': config['name'],
                    'old_quality': old_quality,
                    'old_quality_mark': old_quality_mark
                })
            else:
                print(f"  ✓ Quality maintained or improved")
        else:
            output_cell.value = result['response']
            sources_cell.value = ""
            print(f"  ❌ Error: {result['response']}")
            
            # If old response was good, this is a degradation
            if old_response and 'error' not in old_response.lower() and old_quality != 'bad':
                severity = "HIGH" if old_quality == "good" else "MEDIUM"
                degraded_responses.append({
                    'serial': row_idx,
                    'prompt': str(prompt),
                    'old_response': old_response,
                    'new_response': result['response'],
                    'old_sources': old_sources,
                    'new_sources': "",
                    'reason': "API call failed, old response was successful",
                    'severity': severity,
                    'sheet_name': config['name'],
                    'old_quality': old_quality,
                    'old_quality_mark': old_quality_mark
                })
    
    # Save new results
    wb_new.save(NEW_OUTPUT_FILE)
    print(f"\n  💾 New responses saved to {NEW_OUTPUT_FILE}")
    
    return degraded_responses, processed, successful

# ===== SHAREPOINT UPLOAD FUNCTION =====
def upload_to_sharepoint(file_path):
    """Upload Excel file to SharePoint via Power Automate and return the file URL"""
    
    if not ENABLE_SHAREPOINT_UPLOAD:
        print("\n  ℹ️  SharePoint upload disabled (ENABLE_SHAREPOINT_UPLOAD = False)")
        return None
    
    if not SHAREPOINT_UPLOAD_URL:
        print("\n  ⚠️  SharePoint upload URL not configured")
        return None
    
    if not os.path.exists(file_path):
        print(f"\n  ⚠️  File not found: {file_path}")
        return None
    
    try:
        import base64
        
        print(f"\n  📤 Uploading {os.path.basename(file_path)} to SharePoint...")
        
        # Read and encode file
        with open(file_path, 'rb') as f:
            file_content = base64.b64encode(f.read()).decode('utf-8')
        
        # Prepare payload
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"{timestamp}_{os.path.basename(file_path)}"
        
        payload = {
            "filename": filename,
            "fileContent": file_content,
            "timestamp": timestamp
        }
        
        # Send to Power Automate
        response = requests.post(
            SHAREPOINT_UPLOAD_URL,
            json=payload,
            headers={"Content-Type": "application/json"},
            timeout=30
        )
        
        if response.status_code == 200:
            result = response.json()
            file_path_relative = result.get('fileUrl', '')
            
            # Convert SharePoint path to full URL
            if file_path_relative:
                # Build full SharePoint URL - use proper encoding for spaces
                site_url = "https://fortive.sharepoint.com/sites/FTV-TheFort"
                # URL encode the path (spaces become %20)
                encoded_path = urllib.parse.quote(file_path_relative)
                # Build direct download/view link
                full_url = f"{site_url}{encoded_path}"
                
                print(f"  ✅ Uploaded successfully!")
                print(f"  📁 File: {file_path_relative}")
                print(f"  🔗 SharePoint URL: {full_url}")
                return full_url
            else:
                print(f"  ⚠️  Upload succeeded but no file URL returned")
                return None
        else:
            print(f"  ❌ Upload failed: HTTP {response.status_code}")
            print(f"  Response: {response.text}")
            return None
            
    except Exception as e:
        print(f"\n  ❌ Error uploading to SharePoint: {e}")
        return None

# ===== TEAMS ALERT FUNCTIONS =====
def send_teams_alert(degraded_responses_summary, sharepoint_url=None):
    """Send Microsoft Teams alert for degraded responses via Office 365 Incoming Webhook"""
    
    if not ENABLE_TEAMS_ALERTS:
        print("\n  ℹ️  Teams alerts disabled (ENABLE_TEAMS_ALERTS = False)")
        return
    
    if not TEAMS_WEBHOOK_URL:
        print("\n  ⚠️  Teams webhook URL not configured, skipping Teams alert")
        return
    
    if not degraded_responses_summary:
        print("\n  ✅ No degraded responses, no alert needed")
        return
    
    # Count by severity
    high_count = sum(1 for d in degraded_responses_summary if d['severity'] == 'HIGH')
    medium_count = sum(1 for d in degraded_responses_summary if d['severity'] == 'MEDIUM')
    
    # Build facts for summary section
    facts = [
        {"name": "Total Degraded Responses", "value": str(len(degraded_responses_summary))},
        {"name": "High Severity Issues", "value": f"🔴 {high_count}"},
        {"name": "Medium Severity Issues", "value": f"🟠 {medium_count}"},
        {"name": "Report File", "value": DEGRADED_OUTPUT_FILE}
    ]
    
    # Add SharePoint link if available
    if sharepoint_url:
        facts.append({"name": "📁 SharePoint Report", "value": f"[Open Report]({sharepoint_url})"})
    
    # Build sections for MessageCard
    sections = [
        {
            "activityTitle": "🚨 API Quality Alert",
            "activitySubtitle": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "facts": facts,
            "markdown": True
        }
    ]
    
    # Add top 5 issues as separate sections
    for idx, deg in enumerate(degraded_responses_summary[:5], 1):
        severity_emoji = "🔴" if deg['severity'] == 'HIGH' else "🟠"
        sections.append({
            "activityTitle": f"{severity_emoji} Issue #{idx}: {deg['sheet_name']} - Row {deg['serial']}",
            "facts": [
                {"name": "Prompt", "value": deg['prompt'][:100] + "..."},
                {"name": "Reason", "value": deg['reason']},
                {"name": "Severity", "value": deg['severity']}
            ]
        })
    
    if len(degraded_responses_summary) > 5:
        sections.append({
            "text": f"... and **{len(degraded_responses_summary) - 5} more issues**. Download the full report for details."
        })
    
    # Office 365 MessageCard format
    message_card = {
        "@type": "MessageCard",
        "@context": "https://schema.org/extensions",
        "summary": f"⚠️ {len(degraded_responses_summary)} Degraded API Responses Detected",
        "themeColor": "FF0000" if high_count > 0 else "FFA500",
        "title": f"🚨 Quality Alert: {len(degraded_responses_summary)} Degraded Responses",
        "sections": sections,
        "potentialAction": [
            {
                "@type": "ActionCard",
                "name": "📊 Report Details",
                "actions": [
                    {
                        "@type": "HttpPOST",
                        "name": "Acknowledge",
                        "target": "https://example.com/acknowledge"  # Optional: add your acknowledgement endpoint
                    }
                ]
            }
        ]
    }
    
    try:
        print(f"\n  📤 Attempting to send Teams alert...")
        response = requests.post(TEAMS_WEBHOOK_URL, json=message_card, timeout=10)
        
        print(f"  📡 Webhook response: HTTP {response.status_code}")
        
        if response.status_code in [200, 202]:
            print(f"  ✅ Teams webhook accepted the request")
            print(f"  ℹ️  Note: Check your Power Automate run history if the card doesn't appear")
            print(f"  ℹ️  URL: https://make.powerautomate.com/")
        else:
            print(f"  ⚠️  Unexpected status code: {response.status_code}")
            print(f"  Response: {response.text[:200] if response.text else '(empty)'}")
    except requests.exceptions.Timeout:
        print(f"  ⚠️  Teams webhook timeout (10s) - continuing anyway")
    except requests.exceptions.RequestException as e:
        print(f"  ⚠️  Teams webhook request failed: {type(e).__name__}")
        print(f"  ℹ️  This won't stop the test - check webhook URL and Power Automate setup")

# ===== MAIN FUNCTION =====
def main():
    """Main execution function"""
    print("="*70)
    print("🧪 INTEGRATED TEST: API Testing + Comparison + Teams Alerts")
    print("="*70)
    print(f"📊 Benchmark File: {BENCHMARK_FILE}")
    print(f"📊 New Output File: {NEW_OUTPUT_FILE}")
    print(f"📊 Degraded Report: {DEGRADED_OUTPUT_FILE}")
    if TEST_LIMIT is None:
        print(f"🔢 Rows per sheet: Full run (no limit)")
    else:
        print(f"🔢 Rows per sheet: {TEST_LIMIT}")
    print(f"📋 Sheets to process: {len(SHEET_CONFIGS)}")
    print("="*70)
    
    all_degraded_responses = []
    total_processed = 0
    total_successful = 0
    
    # Process each sheet
    for sheet_number, config in SHEET_CONFIGS.items():
        try:
            degraded, processed, successful = process_sheet_with_comparison(sheet_number, config)
            all_degraded_responses.extend(degraded)
            total_processed += processed
            total_successful += successful
        except Exception as e:
            print(f"\n  ❌ Error processing sheet {sheet_number}: {e}")
            continue
    
    # Create degraded responses report
    if all_degraded_responses:
        print(f"\n{'='*70}")
        print(f"📝 Creating Degraded Responses Report")
        print(f"{'='*70}")
        
        wb_degraded = openpyxl.Workbook()
        wb_degraded.remove(wb_degraded.active)  # Remove default sheet
        
        # Group by sheet
        sheets_with_issues = {}
        for deg in all_degraded_responses:
            sheet_name = deg['sheet_name']
            if sheet_name not in sheets_with_issues:
                sheets_with_issues[sheet_name] = []
            sheets_with_issues[sheet_name].append(deg)
        
        # Create a sheet for each mentor with issues
        for sheet_name, issues in sheets_with_issues.items():
            ws = create_degraded_responses_sheet(wb_degraded, sheet_name)
            
            for issue in issues:
                quality_display = f"{issue.get('old_quality', 'unknown').upper()}"
                if issue.get('old_quality_mark'):
                    quality_display += f" ({issue['old_quality_mark']})"
                
                row_data = [
                    issue['serial'],
                    issue['prompt'],
                    issue['old_response'],
                    issue['new_response'],
                    issue['old_sources'],
                    issue['new_sources'],
                    quality_display,
                    issue['reason'],
                    issue['severity']
                ]
                ws.append(row_data)
                
                # Set text wrapping and row height
                current_row = ws.max_row
                for col_num in range(2, 10):
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                # Auto-adjust row height based on content (estimate)
                max_lines = 1
                for col_num in [2, 3, 4, 8]:  # Prompt, Old Response, New Response, Reason
                    cell = ws.cell(row=current_row, column=col_num)
                    if cell.value:
                        text_length = len(str(cell.value))
                        col_width = ws.column_dimensions[cell.column_letter].width
                        estimated_lines = max(1, text_length / col_width)
                        max_lines = max(max_lines, estimated_lines)
                
                # Set minimum row height with some padding
                ws.row_dimensions[current_row].height = max(30, min(max_lines * 15, 200))
                
                # Color code by severity
                severity_cell = ws.cell(row=current_row, column=9)
                if issue['severity'] == 'HIGH':
                    severity_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                else:
                    severity_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        wb_degraded.save(DEGRADED_OUTPUT_FILE)
        print(f"  💾 Degraded responses report saved: {DEGRADED_OUTPUT_FILE}")
    else:
        # Create a minimal report even when there are no degraded responses so CI/artifact
        # steps that expect the file won't fail. This file will contain a single sheet
        # with a message indicating no degradations were found.
        print("\nNo degraded responses found — creating empty report for records.")
        wb_degraded = openpyxl.Workbook()
        ws = wb_degraded.active
        ws.title = 'Summary'
        ws.append(["Status", "Message"]) 
        ws.append(["OK", "No degraded responses found in this run."])
        # basic styling
        ws.row_dimensions[1].height = 20
        wb_degraded.save(DEGRADED_OUTPUT_FILE)
        print(f"  💾 Empty degraded responses report saved: {DEGRADED_OUTPUT_FILE}")
    
    # Upload to SharePoint
    sharepoint_url = None
    if all_degraded_responses:
        print(f"\n{'='*70}")
        print(f"☁️  Uploading to SharePoint")
        print(f"{'='*70}")
        sharepoint_url = upload_to_sharepoint(DEGRADED_OUTPUT_FILE)
    
    # Send Teams alert
    print(f"\n{'='*70}")
    print(f"📢 Sending Teams Alert")
    print(f"{'='*70}")
    send_teams_alert(all_degraded_responses, sharepoint_url)
    
    # Final summary
    print(f"\n{'='*70}")
    print("🎉 TEST COMPLETE!")
    print(f"{'='*70}")
    print(f"📊 Sheets processed: {len(SHEET_CONFIGS)}")
    print(f"📝 Total rows processed: {total_processed}")
    print(f"✅ Successful API calls: {total_successful}")
    print(f"❌ Failed API calls: {total_processed - total_successful}")
    print(f"⚠️  Degraded responses: {len(all_degraded_responses)}")
    if all_degraded_responses:
        high_count = sum(1 for d in all_degraded_responses if d['severity'] == 'HIGH')
        medium_count = sum(1 for d in all_degraded_responses if d['severity'] == 'MEDIUM')
        print(f"   🔴 High severity: {high_count}")
        print(f"   🟠 Medium severity: {medium_count}")
    print(f"💾 New responses: {NEW_OUTPUT_FILE}")
    if all_degraded_responses:
        print(f"💾 Degraded report: {DEGRADED_OUTPUT_FILE}")
    print("="*70)

if __name__ == "__main__":
    main()
