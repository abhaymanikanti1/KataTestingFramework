"""
KATA Workflow Search Functionality Test
Tests the KATA workflow search API endpoint locally
Stores results in an Excel file for analysis
"""
import requests
import urllib.parse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import json
import uuid

# ===== CONFIG =====
KATA_WORKFLOW_API_URL = "https://kata-workflow-d7hncaeta9cra7a8.eastus-01.azurewebsites.net/api/kata_workflow"
EMAIL_ID = "abhay.manikanti@fortive.com"
API_KEY = "d7e8f9b6-92a4-48e2-a0cd-f81c993f29c1"
SESSION_ID = "43908e3d-7fee-4688-a6c5-f3bd32a94ffd"

# Try multiple header configurations
HEADERS_VARIANTS = [
    # Variant 1: Standard form-urlencoded
    {
        'Content-Type': "application/x-www-form-urlencoded",
        'User-Agent': "insomnia/11.4.0"
    },
    # Variant 2: With API Key header
    {
        'Content-Type': "application/x-www-form-urlencoded",
        'User-Agent': "insomnia/11.4.0",
        'X-API-Key': API_KEY
    },
    # Variant 3: With Authorization header
    {
        'Content-Type': "application/x-www-form-urlencoded",
        'User-Agent': "insomnia/11.4.0",
        'Authorization': f'Bearer {API_KEY}'
    },
    # Variant 4: JSON content type
    {
        'Content-Type': "application/json",
        'User-Agent': "insomnia/11.4.0",
        'X-API-Key': API_KEY
    },
    # Variant 5: JSON with Bearer token
    {
        'Content-Type': "application/json",
        'User-Agent': "insomnia/11.4.0",
        'Authorization': f'Bearer {API_KEY}'
    }
]

HEADERS = HEADERS_VARIANTS[0]  # Default headers

# Excel file to store results
OUTPUT_FILE = "kata_workflow_search_results.xlsx"

# Test queries - add your search queries here
TEST_QUERIES = [
    "What is PSP?",
    "Explain VSM methodology",
    "What is TPI in quality management?",
    "How to implement continuous improvement?",
    "What are the key principles of lean manufacturing?",
    "Define process improvement",
    "What is KATA workflow?",
    "How to use KATA for problem solving?",
    "Explain DIVE framework",
    "What are improvement cycles?"
]


def test_kata_workflow_search(query):
    """
    Test the KATA workflow search API with a given query
    
    Args:
        query (str): Search query to test
        
    Returns:
        dict: Response containing status, answer, sources, and metadata
    """
    conversation_id = str(uuid.uuid4())
    
    # Try different payload structures
    payloads_to_try = [
        # Standard format from integrated_test_comparison.py
        {
            "email_id": EMAIL_ID,
            "question": query,
            "session_id": SESSION_ID,
            "conversation_id": conversation_id,
            "agent_id": "search",
            "thread_id": "",
            "selected_column": "",
            "container": "useruploaded"
        },
        # Simplified format
        {
            "email_id": EMAIL_ID,
            "question": query,
            "session_id": SESSION_ID,
        },
        # With explicit search context
        {
            "email_id": EMAIL_ID,
            "query": query,
            "session_id": SESSION_ID,
            "conversation_id": conversation_id,
        }
    ]
    
    for idx, payload_dict in enumerate(payloads_to_try, 1):
        # Try each payload with different header variants
        for headers_idx, headers in enumerate(HEADERS_VARIANTS, 1):
            payload_encoded = urllib.parse.urlencode(payload_dict)
            
            # For JSON content type, send JSON instead of form-urlencoded
            if headers.get('Content-Type') == 'application/json':
                request_data = json.dumps(payload_dict)
            else:
                request_data = payload_encoded
            
            try:
                print(f"  Attempt {idx}.{headers_idx}: Testing payload variant {idx}, headers variant {headers_idx}...")
                response = requests.post(
                    KATA_WORKFLOW_API_URL,
                    data=request_data,
                    headers=headers,
                    verify=False,
                    timeout=30
                )
                
                print(f"  Status Code: {response.status_code}")
                
                if response.status_code == 200:
                    # Successfully got response
                    response_text = response.text.strip()
                    
                    # Parse response
                    if response_text:
                        try:
                            data = json.loads(response_text)
                            return {
                                'status': 'success',
                                'status_code': 200,
                                'response': data.get('answer', data.get('response', str(data))),
                                'sources': data.get('sources', []),
                                'metadata': data,
                                'payload_variant': idx,
                                'headers_variant': headers_idx
                            }
                        except json.JSONDecodeError:
                            return {
                                'status': 'success',
                                'status_code': 200,
                                'response': response_text,
                                'sources': [],
                                'metadata': {'raw_text': response_text},
                                'payload_variant': idx,
                                'headers_variant': headers_idx
                            }
                    else:
                        return {
                            'status': 'success',
                            'status_code': 200,
                            'response': 'Empty response',
                            'sources': [],
                            'metadata': {},
                            'payload_variant': idx,
                            'headers_variant': headers_idx
                        }
                
                elif response.status_code == 401:
                    print(f"  ⚠️ Unauthorized - trying next variant...")
                    continue
                    
                else:
                    # Non-200, non-401 status - return error
                    return {
                        'status': 'error',
                        'status_code': response.status_code,
                        'response': f"HTTP {response.status_code}: {response.text[:200]}",
                        'sources': [],
                        'metadata': {'error_detail': response.text},
                        'payload_variant': idx,
                        'headers_variant': headers_idx
                    }
                    
            except requests.Timeout:
                print(f"  ⏱️ Request timeout")
                continue
            except Exception as e:
                print(f"  ❌ Exception: {str(e)}")
                continue
    
    # All attempts failed
    return {
        'status': 'error',
        'status_code': 401,
        'response': 'All authentication attempts failed (401 Unauthorized)',
        'sources': [],
        'metadata': {},
        'payload_variant': None
    }


def create_excel_report(results):
    """
    Create an Excel file with test results
    
    Args:
        results (list): List of test result dictionaries
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KATA Workflow Search Tests"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    success_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    error_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["#", "Query", "Status", "Status Code", "Response", "Sources", "Payload Variant", "Headers Variant", "Timestamp"]
    ws.append(headers)
    
    # Style headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Add data
    for idx, result in enumerate(results, 1):
        sources_text = "\n".join(result.get('sources', [])) if result.get('sources') else "No sources"
        
        row_data = [
            idx,
            result['query'],
            result['status'],
            result.get('status_code', 'N/A'),
            result['response'],
            sources_text,
            result.get('payload_variant', 'N/A'),
            result.get('headers_variant', 'N/A'),
            result['timestamp']
        ]
        
        ws.append(row_data)
        
        # Style row based on status
        row_num = idx + 1
        fill = success_fill if result['status'] == 'success' else error_fill
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = fill
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = border
    
    # Adjust column widths
    column_widths = [5, 40, 10, 12, 80, 60, 15, 15, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add summary sheet
    summary_ws = wb.create_sheet("Summary", 0)
    summary_ws["A1"] = "KATA Workflow Search Test Summary"
    summary_ws["A1"].font = Font(bold=True, size=14)
    
    summary_ws["A3"] = "Total Tests:"
    summary_ws["B3"] = len(results)
    
    successful = sum(1 for r in results if r['status'] == 'success')
    failed = len(results) - successful
    
    summary_ws["A4"] = "Successful:"
    summary_ws["B4"] = successful
    summary_ws["B4"].fill = success_fill
    
    summary_ws["A5"] = "Failed:"
    summary_ws["B5"] = failed
    summary_ws["B5"].fill = error_fill
    
    summary_ws["A6"] = "Success Rate:"
    summary_ws["B6"] = f"{(successful/len(results)*100):.1f}%" if results else "N/A"
    
    summary_ws["A8"] = "Test Date:"
    summary_ws["B8"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    summary_ws["A9"] = "API Endpoint:"
    summary_ws["B9"] = KATA_WORKFLOW_API_URL
    summary_ws.column_dimensions["B"].width = 80
    
    # Save workbook
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Results saved to: {OUTPUT_FILE}")


def main():
    """Main test execution function"""
    print("=" * 80)
    print("KATA Workflow Search Functionality Test")
    print("=" * 80)
    print(f"API Endpoint: {KATA_WORKFLOW_API_URL}")
    print(f"Test Queries: {len(TEST_QUERIES)}")
    print(f"Output File: {OUTPUT_FILE}")
    print("=" * 80)
    print()
    
    results = []
    
    for idx, query in enumerate(TEST_QUERIES, 1):
        print(f"[{idx}/{len(TEST_QUERIES)}] Testing: {query}")
        
        # Test the query
        result = test_kata_workflow_search(query)
        
        # Add metadata
        result['query'] = query
        result['timestamp'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        results.append(result)
        
        # Print result summary
        if result['status'] == 'success':
            print(f"  ✅ Success! Response: {result['response'][:100]}...")
        else:
            print(f"  ❌ Failed: {result['response']}")
        
        print()
    
    # Create Excel report
    print("\nGenerating Excel report...")
    create_excel_report(results)
    
    # Print summary
    successful = sum(1 for r in results if r['status'] == 'success')
    failed = len(results) - successful
    
    print("\n" + "=" * 80)
    print("TEST SUMMARY")
    print("=" * 80)
    print(f"Total Tests: {len(results)}")
    print(f"✅ Successful: {successful}")
    print(f"❌ Failed: {failed}")
    print(f"Success Rate: {(successful/len(results)*100):.1f}%")
    print("=" * 80)


if __name__ == "__main__":
    main()
