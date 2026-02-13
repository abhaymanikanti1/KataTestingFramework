"""
FastAPI Server for Degraded Responses Data
Serves Excel data from Azure Blob Storage as JSON for frontend consumption

Usage:
    python api_server.py
    
    Or with uvicorn:
    uvicorn api_server:app --host 0.0.0.0 --port 8000 --reload

API Endpoints:
    GET  /                                      - API info
    GET  /api/health                            - Health check
    GET  /api/degraded-responses                - All degraded responses
    GET  /api/degraded-responses/{sheet_name}   - Specific sheet data
    POST /api/refresh                           - Force cache refresh
    GET  /docs                                  - Interactive API documentation
"""
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
import openpyxl
from datetime import datetime
import os
from typing import List, Dict, Any, Optional
from blob_storage_uploader import download_excel_from_blob, BLOB_NAME, check_blob_exists, get_blob_metadata

app = FastAPI(
    title="KATA Testing API",
    description="API for serving degraded API responses data from Azure Blob Storage",
    version="1.0.0",
    docs_url="/docs",
    redoc_url="/redoc"
)

# Enable CORS for frontend access
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In production, specify your frontend domain like ["https://yourfrontend.com"]
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Cache for Excel data (refresh periodically)
data_cache = {
    'data': None,
    'last_updated': None,
    'cache_duration': 300  # 5 minutes in seconds
}

def read_excel_from_blob() -> Optional[List[Dict[str, Any]]]:
    """
    Download Excel from Blob Storage and parse into JSON format
    
    Returns:
        List of dictionaries containing sheet data, or None if failed
    """
    try:
        print(f"[{datetime.now()}] Downloading Excel from Blob Storage...")
        
        # Download from Blob Storage
        local_file = download_excel_from_blob()
        
        if not local_file or not os.path.exists(local_file):
            print(f"[{datetime.now()}] Failed to download file from Blob Storage")
            return None
        
        print(f"[{datetime.now()}] Parsing Excel file...")
        
        # Parse Excel
        wb = openpyxl.load_workbook(local_file, data_only=True)
        
        all_sheets_data = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Get headers (first row)
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value))
            
            if not headers:
                print(f"[{datetime.now()}] Warning: Sheet '{sheet_name}' has no headers, skipping")
                continue
            
            # Get data rows
            sheet_data = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_dict = {}
                has_data = False
                
                for idx, value in enumerate(row):
                    if idx < len(headers):
                        # Convert cell values to JSON-serializable types
                        if value is None:
                            row_dict[headers[idx]] = None
                        elif isinstance(value, datetime):
                            row_dict[headers[idx]] = value.isoformat()
                            has_data = True
                        else:
                            row_dict[headers[idx]] = str(value)
                            has_data = True
                
                # Only add row if it has data
                if has_data:
                    sheet_data.append(row_dict)
            
            all_sheets_data.append({
                'sheet_name': sheet_name,
                'row_count': len(sheet_data),
                'data': sheet_data
            })
            
            print(f"[{datetime.now()}] Parsed sheet '{sheet_name}': {len(sheet_data)} rows")
        
        # Clean up temp file
        try:
            os.remove(local_file)
            print(f"[{datetime.now()}] Cleaned up temp file")
        except:
            pass
        
        print(f"[{datetime.now()}] Successfully parsed {len(all_sheets_data)} sheets")
        return all_sheets_data
        
    except Exception as e:
        print(f"[{datetime.now()}] Error reading Excel: {e}")
        import traceback
        traceback.print_exc()
        return None

def get_cached_data() -> Optional[List[Dict[str, Any]]]:
    """
    Get data from cache or refresh if expired
    
    Returns:
        Cached data or freshly loaded data, None if unable to load
    """
    now = datetime.now()
    
    # Check if cache is valid
    if data_cache['data'] is not None and data_cache['last_updated'] is not None:
        elapsed = (now - data_cache['last_updated']).total_seconds()
        if elapsed < data_cache['cache_duration']:
            print(f"[{now}] Returning cached data (age: {elapsed:.0f}s)")
            return data_cache['data']
    
    # Refresh cache
    print(f"[{now}] Cache expired or empty, refreshing from Blob Storage...")
    data = read_excel_from_blob()
    
    if data is not None:
        data_cache['data'] = data
        data_cache['last_updated'] = now
        print(f"[{now}] ✅ Cache updated successfully")
    else:
        print(f"[{now}] ⚠️ Failed to refresh cache")
    
    return data

@app.get("/", tags=["Info"])
def root():
    """
    API root - provides information about available endpoints
    """
    return {
        "status": "ok",
        "service": "KATA Testing API",
        "version": "1.0.0",
        "description": "Serves degraded API responses data from Azure Blob Storage",
        "endpoints": {
            "info": "/",
            "health": "/api/health",
            "all_data": "/api/degraded-responses",
            "sheet_data": "/api/degraded-responses/{sheet_name}",
            "refresh": "/api/refresh",
            "docs": "/docs",
            "redoc": "/redoc"
        }
    }

@app.get("/api/health", tags=["Health"])
def health_check():
    """
    Health check endpoint - returns service status and cache information
    """
    blob_exists = check_blob_exists()
    metadata = get_blob_metadata() if blob_exists else None
    
    cache_age = None
    if data_cache['last_updated']:
        cache_age = (datetime.now() - data_cache['last_updated']).total_seconds()
    
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "cache": {
            "status": "loaded" if data_cache['data'] is not None else "empty",
            "last_updated": data_cache['last_updated'].isoformat() if data_cache['last_updated'] else None,
            "age_seconds": cache_age,
            "ttl_seconds": data_cache['cache_duration']
        },
        "blob_storage": {
            "connected": blob_exists,
            "blob_exists": blob_exists,
            "blob_metadata": metadata
        }
    }

@app.get("/api/degraded-responses", tags=["Data"])
def get_all_degraded_responses():
    """
    Get all degraded responses from all sheets
    
    Returns:
        JSON containing all sheets with degraded responses data
        
    Response format:
    ```json
    {
        "total_sheets": 4,
        "total_issues": 15,
        "last_updated": "2026-02-12T10:30:00",
        "sheets": [
            {
                "sheet_name": "PSP Mentor",
                "row_count": 5,
                "data": [
                    {
                        "Serial Number": "5",
                        "Prompt": "What is DIVE?",
                        "Old Response (Benchmark)": "...",
                        "New Response": "...",
                        "Old Sources": "...",
                        "New Sources": "...",
                        "Benchmark Quality": "GOOD",
                        "Degradation Reason": "Response 60% shorter",
                        "Severity": "HIGH"
                    }
                ]
            }
        ]
    }
    ```
    """
    data = get_cached_data()
    
    if data is None:
        raise HTTPException(
            status_code=503, 
            detail="Unable to load data from Blob Storage. Check if the Excel file exists and the connection string is configured."
        )
    
    total_issues = sum(sheet['row_count'] for sheet in data)
    
    return {
        "total_sheets": len(data),
        "total_issues": total_issues,
        "last_updated": data_cache['last_updated'].isoformat() if data_cache['last_updated'] else None,
        "sheets": data
    }

@app.get("/api/degraded-responses/{sheet_name}", tags=["Data"])
def get_degraded_responses_by_sheet(sheet_name: str):
    """
    Get degraded responses for a specific sheet
    
    Args:
        sheet_name: One of "PSP Mentor", "VSM Mentor", "TPI Mentor", "Search/Chat"
        
    Returns:
        JSON containing data for the specified sheet
        
    Example:
        GET /api/degraded-responses/PSP%20Mentor
    """
    data = get_cached_data()
    
    if data is None:
        raise HTTPException(
            status_code=503, 
            detail="Unable to load data from Blob Storage"
        )
    
    # Find matching sheet (case-insensitive, handle URL encoding)
    sheet_name_normalized = sheet_name.lower().replace('%20', ' ')
    
    for sheet in data:
        if sheet['sheet_name'].lower() == sheet_name_normalized:
            return {
                "sheet_name": sheet['sheet_name'],
                "total_issues": sheet['row_count'],
                "last_updated": data_cache['last_updated'].isoformat() if data_cache['last_updated'] else None,
                "data": sheet['data']
            }
    
    # List available sheets in error message
    available_sheets = [sheet['sheet_name'] for sheet in data]
    raise HTTPException(
        status_code=404, 
        detail=f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(available_sheets)}"
    )

@app.post("/api/refresh", tags=["Admin"])
def force_refresh():
    """
    Force refresh data from Blob Storage (manual cache invalidation)
    
    Useful when you've uploaded a new Excel file and want to see changes immediately
    without waiting for cache to expire.
    
    Returns:
        Status of the refresh operation
    """
    print(f"[{datetime.now()}] Manual refresh requested")
    
    # Clear cache
    data_cache['data'] = None
    data_cache['last_updated'] = None
    
    # Force reload
    data = get_cached_data()
    
    if data is None:
        raise HTTPException(
            status_code=503, 
            detail="Unable to refresh data from Blob Storage"
        )
    
    total_issues = sum(sheet['row_count'] for sheet in data)
    
    return {
        "status": "success",
        "message": "Data refreshed from Blob Storage",
        "refreshed_at": datetime.now().isoformat(),
        "total_sheets": len(data),
        "total_issues": total_issues,
        "sheets": [{"name": sheet['sheet_name'], "rows": sheet['row_count']} for sheet in data]
    }

@app.get("/api/sheets", tags=["Info"])
def list_sheets():
    """
    List all available sheets without returning full data
    
    Returns:
        List of sheet names and row counts
    """
    data = get_cached_data()
    
    if data is None:
        raise HTTPException(
            status_code=503, 
            detail="Unable to load data from Blob Storage"
        )
    
    return {
        "total_sheets": len(data),
        "sheets": [
            {
                "name": sheet['sheet_name'],
                "row_count": sheet['row_count']
            } 
            for sheet in data
        ]
    }

if __name__ == "__main__":
    import uvicorn
    
    print("="*70)
    print("🚀 Starting KATA Testing API Server")
    print("="*70)
    print(f"📡 Server: http://0.0.0.0:8000")
    print(f"📚 API Docs: http://0.0.0.0:8000/docs")
    print(f"📖 ReDoc: http://0.0.0.0:8000/redoc")
    print("="*70)
    
    uvicorn.run(
        app, 
        host="0.0.0.0", 
        port=8000,
        log_level="info"
    )
