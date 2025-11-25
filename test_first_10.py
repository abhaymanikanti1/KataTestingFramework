"""
Test script: Process first 10 questions from all sheets (PSP, VSM, TPI, Search)
This tests the API integration before full automation.
"""
import openpyxl
import requests
import uuid
from datetime import datetime
import urllib.parse
import os
import re
import json
from openpyxl.styles import Font, Alignment, PatternFill

# ===== CONFIG =====
EXCEL_FILE = "Direct Query Master List V3.xlsx"

def get_next_version_filename():
    """Generate the next version filename starting from V3"""
    base_name = "Direct Query Master List"
    current_dir = "/Users/abhay.manikanti/Documents/untitled folder"
    
    existing_versions = []
    for filename in os.listdir(current_dir):
        if filename.startswith(base_name) and filename.endswith('.xlsx'):
            match = re.search(r'V(\d+)', filename)
            if match:
                version_num = int(match.group(1))
                existing_versions.append(version_num)
    
    if existing_versions:
        next_version = max(existing_versions) + 1
    else:
        next_version = 3
    
    next_version = max(next_version, 3)
    
    return f"{base_name} V{next_version}.xlsx"

OUTPUT_FILE = get_next_version_filename()

# --- Base API Configuration ---
API_BASE_URL = "https://container-app-ui-stage.purplesky-e0183d2f.eastus.azurecontainerapps.io/"
EMAIL_ID = "abhay.manikanti@fortive.com"
API_KEY = "d7e8f9b6-92a4-48e2-a0cd-f81c993f29c1"

HEADERS = {
    'Content-Type': "application/x-www-form-urlencoded",
    'User-Agent': "insomnia/11.4.0"
}

# Test Mode: Process only first 10 rows
TEST_LIMIT = 10

# Sheet configurations (sheet number -> API config)
SHEET_CONFIGS = {
    3: {
        'name': 'PSP Mentor',
        'api_path': '/api/pspmentor',
        'agent_id': 'psp'
    },
    4: {
        'name': 'VSM Mentor',
        'api_path': '/api/vsmmentor',
        'agent_id': 'vsm'
    },
    5: {
        'name': 'TPI Mentor',
        'api_path': '/api/tpimentor',
        'agent_id': 'tpi'
    },
    6: {
        'name': 'Search/Chat',
        'api_path': '/api/chat',
        'agent_id': 'search'
    }
}

# Session ID
SESSION_ID = "43908e3d-7fee-4688-a6c5-f3bd32a94ffd"

# ===== FUNCTIONS =====
def send_question_to_api(prompt, api_url, agent_id):
    """Send a question to the API and return the response"""
    conversation_id = str(uuid.uuid4())
    
    payload_dict = {
        "email_id": EMAIL_ID,
        "question": prompt,
        "session_id": SESSION_ID,
        "conversation_id": conversation_id,
        "agent_id": agent_id,
        "thread_id": "",
        "selected_column": "",
        "container": "useruploaded"
    }
    
    payload_encoded = urllib.parse.urlencode(payload_dict)
    
    try:
        response = requests.post(
            api_url, 
            data=payload_encoded, 
            headers=HEADERS, 
            verify=False, 
            timeout=60
        )
        
        if response.status_code == 200:
            if not response.text.strip():
                return {
                    'status': 'success',
                    'response': 'Empty response from API',
                    'sources': []
                }
            
            response_text = response.text.strip()
            
            # Check if this is SSE format
            if response_text.startswith('data:') or '\ndata:' in response_text:
                return parse_sse_response(response_text)
            else:
                return parse_json_response(response_text)
        else:
            return {
                'status': 'error',
                'response': f"HTTP {response.status_code}",
                'sources': []
            }
            
    except requests.Timeout:
        return {
            'status': 'error',
            'response': 'Request timeout (60s)',
            'sources': []
        }
    except Exception as e:
        return {
            'status': 'error',
            'response': f"Error: {str(e)[:100]}",
            'sources': []
        }

def parse_sse_response(response_text):
    """Parse Server-Sent Events (SSE) format response"""
    content_parts = []
    extracted_urls = []
    
    for line in response_text.split('\n'):
        line = line.strip()
        if line.startswith('data: '):
            try:
                json_part = line[6:].strip()
                if json_part and json_part != '[DONE]':
                    data_obj = json.loads(json_part)
                    
                    if isinstance(data_obj, dict):
                        # NEW FORMAT: assistant_output directly in the object
                        if 'assistant_output' in data_obj and data_obj['assistant_output']:
                            output = str(data_obj['assistant_output'])
                            # Filter out initialization messages
                            if not any(emoji in output for emoji in ['🔄', '🔧', '📋', '🔍', '📂', '🤔', '📚', '✨']):
                                content_parts.append(output)
                        
                        # OLD FORMAT: nested under 'data' key
                        elif 'data' in data_obj and isinstance(data_obj['data'], dict):
                            chunk_data = data_obj['data']
                            
                            # Extract content
                            if 'assistant_output' in chunk_data and chunk_data['assistant_output']:
                                content_parts.append(str(chunk_data['assistant_output']))
                            elif 'content' in chunk_data and chunk_data['content']:
                                content_parts.append(str(chunk_data['content']))
                            
                            # Extract sources
                            if 'sources' in chunk_data and isinstance(chunk_data['sources'], list):
                                for source in chunk_data['sources']:
                                    if isinstance(source, dict):
                                        for url_field in ['page_url', 'url', 'link', 'source_url']:
                                            if url_field in source and source[url_field]:
                                                extracted_urls.append(source[url_field])
                                                break
                        
                        # Fallback: direct content field
                        elif 'content' in data_obj and data_obj['content']:
                            content_parts.append(str(data_obj['content']))
                        
                        # Extract sources from top-level sources field
                        if 'sources' in data_obj and isinstance(data_obj['sources'], list):
                            for source in data_obj['sources']:
                                if isinstance(source, dict):
                                    for url_field in ['page_url', 'url', 'link', 'source_url']:
                                        if url_field in source and source[url_field]:
                                            extracted_urls.append(source[url_field])
                                            break
                            
            except json.JSONDecodeError:
                continue
    
    text_content = ''.join(content_parts)
    
    if not text_content:
        text_content = "SSE response: No content extracted"
    
    # Extract URLs from content if no sources found
    if not extracted_urls and text_content:
        url_pattern = r'https?://[^\s<>"{}|\\^`\[\]]+'
        content_urls = re.findall(url_pattern, text_content + response_text)
        extracted_urls.extend(content_urls[:5])
    
    unique_urls = list(dict.fromkeys(extracted_urls))
    
    return {
        'status': 'success',
        'response': text_content,
        'sources': unique_urls[:10]
    }

def parse_json_response(response_text):
    """Parse regular JSON format response"""
    try:
        result = json.loads(response_text)
        
        text_content = ""
        if isinstance(result, dict) and "content" in result:
            text_content = str(result["content"])
        else:
            text_content = "Unexpected JSON format: " + str(result)[:200]
        
        return {
            'status': 'success',
            'response': text_content,
            'sources': []
        }
    except json.JSONDecodeError as e:
        return {
            'status': 'error',
            'response': f"JSON parse error: {str(e)}",
            'sources': []
        }

def process_sheet(wb, sheet_number, config):
    """Process first 10 rows of a specific sheet"""
    sheetnames = wb.sheetnames
    
    if sheet_number < 1 or sheet_number > len(sheetnames):
        print(f"  ❌ Invalid sheet number: {sheet_number}")
        return 0, 0
    
    ws = wb[sheetnames[sheet_number - 1]]
    sheet_name = config['name']
    api_url = f"{API_BASE_URL}{config['api_path']}"
    agent_id = config['agent_id']
    
    print(f"\n{'='*70}")
    print(f"📊 Processing Sheet {sheet_number}: {sheet_name}")
    print(f"🌐 API: {api_url}")
    print(f"🤖 Agent: {agent_id}")
    print(f"⚠️  TEST MODE: First {TEST_LIMIT} rows only")
    print(f"{'='*70}")
    
    processed = 0
    successful = 0
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 1):
        # Stop after TEST_LIMIT rows
        if processed >= TEST_LIMIT:
            print(f"\n  ⏸️  Reached test limit of {TEST_LIMIT} rows")
            break
        
        serial_cell = row[0]   # Column A: Serial Number
        prompt_cell = row[1]   # Column B: Prompt
        output_cell = row[2]   # Column C: Output
        sources_cell = row[3]  # Column D: Sources
        
        prompt = prompt_cell.value
        if not prompt or not str(prompt).strip():
            continue  # Skip empty rows
        
        processed += 1
        print(f"\n[{processed}/{TEST_LIMIT}] Row {row_idx}:")
        print(f"  📝 Prompt: {str(prompt)[:60]}...")
        
        # Send to API
        result = send_question_to_api(str(prompt), api_url, agent_id)
        
        if result['status'] == 'success':
            output_cell.value = result['response']
            sources_cell.value = "\n".join(result['sources']) if result['sources'] else ""
            successful += 1
            print(f"  ✅ Success: {len(result['response'])} chars")
            if result['sources']:
                print(f"  📚 Sources: {len(result['sources'])} URLs")
        else:
            output_cell.value = result['response']
            sources_cell.value = ""
            print(f"  ❌ Error: {result['response']}")
    
    return processed, successful

def main():
    """Main execution function"""
    print("="*70)
    print("🧪 TEST MODE: First 10 Rows - All Sheets")
    print("="*70)
    print(f"📊 Input Excel: {EXCEL_FILE}")
    print(f"💾 Output Excel: {OUTPUT_FILE}")
    print(f"🔢 Rows per sheet: {TEST_LIMIT}")
    print(f"📋 Sheets to process: {len(SHEET_CONFIGS)}")
    print("="*70)
    
    # Load workbook
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        print(f"\n✅ Loaded workbook with {len(wb.sheetnames)} sheets")
    except FileNotFoundError:
        print(f"\n❌ Error: Excel file not found: {EXCEL_FILE}")
        return
    except Exception as e:
        print(f"\n❌ Error loading workbook: {e}")
        return
    
    # Process each configured sheet
    total_processed = 0
    total_successful = 0
    
    for sheet_number, config in SHEET_CONFIGS.items():
        try:
            processed, successful = process_sheet(wb, sheet_number, config)
            total_processed += processed
            total_successful += successful
            
            # Save after each sheet
            wb.save(OUTPUT_FILE)
            print(f"\n  💾 Progress saved to {OUTPUT_FILE}")
            
        except Exception as e:
            print(f"\n  ❌ Error processing sheet {sheet_number}: {e}")
            continue
    
    # Final summary
    print("\n" + "="*70)
    print("🎉 TEST COMPLETE!")
    print("="*70)
    print(f"📊 Sheets processed: {len(SHEET_CONFIGS)}")
    print(f"📝 Total rows processed: {total_processed}")
    print(f"✅ Successful API calls: {total_successful}")
    print(f"❌ Failed API calls: {total_processed - total_successful}")
    print(f"💾 Results saved to: {OUTPUT_FILE}")
    print("="*70)
    
    # Azure deployment recommendation
    print("\n" + "="*70)
    print("☁️  AZURE DEPLOYMENT RECOMMENDATIONS")
    print("="*70)
    print("""
For daily automation, consider these options:

1. ✅ AZURE FUNCTIONS (Recommended for this use case)
   Pros:
   - Serverless, pay-per-execution
   - Built-in timer triggers (run daily automatically)
   - No infrastructure management
   - Auto-scaling
   - Cost-effective for periodic tasks
   - Easy to deploy and update
   
   Cons:
   - 10-minute timeout (Premium plan for longer)
   - Cold start delays
   
   Best for: Scheduled daily tasks with short runtime

2. ⚡ AZURE CONTAINER INSTANCES
   Pros:
   - Full control over environment
   - No timeout limits
   - Can handle long-running processes
   - Custom dependencies easily
   
   Cons:
   - More expensive (always running or manual start/stop)
   - Requires more setup and management
   - Need to implement scheduling separately
   
   Best for: Long-running jobs or complex dependencies

3. 🔄 AZURE LOGIC APPS + FUNCTIONS
   Pros:
   - Visual workflow designer
   - Built-in schedule triggers
   - Easy integration with other services
   - Error handling and retry logic
   
   Cons:
   - Can be more expensive
   - Learning curve for Logic Apps
   
   Best for: Complex workflows with multiple steps

💡 RECOMMENDATION:
For your daily automation processing 4 sheets with ~hundreds of rows:
→ Use AZURE FUNCTIONS with Timer Trigger (CRON: 0 0 9 * * *)
   - Run daily at 9 AM
   - Consumption plan initially (upgrade to Premium if needed)
   - Store Excel in Azure Blob Storage
   - Use Application Insights for monitoring

Estimated cost: $5-20/month depending on execution time
""")
    print("="*70)

if __name__ == "__main__":
    main()
