"""
End-to-end test: Create Excel → Upload to SharePoint → Send Teams Alert with link
"""
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

sys.path.insert(0, '/Users/abhay.manikanti/Documents/untitled folder')

from integrated_test_comparison import upload_to_sharepoint, send_teams_alert

print('🧪 End-to-End SharePoint + Teams Integration Test\n')

# Step 1: Create a test Excel file
print('📝 Step 1: Creating test Excel file...')
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Degraded Responses"

# Headers
headers = ['Sheet', 'Serial', 'Prompt', 'Response', 'Benchmark', 'Reason', 'Severity']
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

# Add test data
test_data = [
    ['PSP Mentor', 5, 'What is DIVE?', 'Short response...', 'Long detailed response...', 'Response 60% shorter', 'HIGH'],
    ['VSM Mentor', 12, 'Explain VSM', 'Generic answer', 'Specific detailed answer', 'Generic content detected', 'MEDIUM'],
    ['TPI Mentor', 8, 'TPI practices', 'Brief answer', 'Comprehensive answer', 'Missing key keywords', 'MEDIUM']
]

for row_data in test_data:
    ws.append(row_data)

test_file = 'TEST_Degraded_Responses_Report.xlsx'
wb.save(test_file)
print(f'  ✅ Created: {test_file}\n')

# Step 2: Upload to SharePoint
print('☁️  Step 2: Uploading to SharePoint...')
sharepoint_url = upload_to_sharepoint(test_file)

if not sharepoint_url:
    print('\n❌ SharePoint upload failed!')
    sys.exit(1)

print(f'\n  ✅ SharePoint upload successful!')
print(f'  🔗 URL: {sharepoint_url}\n')

# Step 3: Send Teams alert
print('📢 Step 3: Sending Teams alert with SharePoint link...')
fake_degraded_responses = [
    {
        'sheet_name': 'PSP Mentor',
        'serial': 5,
        'prompt': 'What is the DIVE process in PSP?',
        'reason': 'Response is 60% shorter than benchmark',
        'severity': 'HIGH'
    },
    {
        'sheet_name': 'VSM Mentor',
        'serial': 12,
        'prompt': 'Explain value stream mapping',
        'reason': 'New response contains generic content',
        'severity': 'MEDIUM'
    },
    {
        'sheet_name': 'TPI Mentor',
        'serial': 8,
        'prompt': 'What are TPI best practices?',
        'reason': 'Missing key keywords from original response',
        'severity': 'MEDIUM'
    }
]

send_teams_alert(fake_degraded_responses, sharepoint_url)

print('\n' + '='*60)
print('✅ END-TO-END TEST COMPLETE!')
print('='*60)
print('\n📋 What to check:')
print('1. SharePoint folder has the Excel file')
print('2. Teams "Kata Bugs Webhook Channel" has the alert card')
print('3. The alert card includes a clickable SharePoint link')
print('\nSharePoint folder:')
print('https://fortive.sharepoint.com/:f:/r/sites/FTV-TheFort/Shared%20Documents/_2025/Projects/KATA/KATA%20Bugs%20Testing%20Automation\n')
